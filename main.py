import os
import telebot
import openpyxl
import json
import requests
import time
import threading
from datetime import datetime,timedelta

API_KEY = ""
chat_id = ""
my_chat_id = ""
with open('config.env', 'r') as openfile:
    json_object = json.load(openfile)
    API_KEY = json_object["API_KEY"]
    chat_id = json_object["CHAT_ID"]
    my_chat_id = json_object["MY_CHAT_ID"]

telegram_poll_url = 'https://api.telegram.org/bot'+API_KEY+'/sendPoll'
telegram_message_url = 'https://api.telegram.org/bot'+API_KEY+'/sendMessage'
telegram_set_commands_url = 'https://api.telegram.org/bot'+API_KEY+'/setMyCommands'

def sendPoll(date_to_post):
  questions = openpyxl.load_workbook("questions.xlsx")
  sheet_obj = questions.active
  total_ques = sheet_obj.max_row

  current_row = 2
  with open('data.json', 'r') as openfile:
    json_object = json.load(openfile)
    current_row = json_object["last_row"]

  ques = sheet_obj.cell(row = current_row, column = 2)
  op1 = sheet_obj.cell(row = current_row, column = 3)
  op2 = sheet_obj.cell(row = current_row, column = 4)
  op3 = sheet_obj.cell(row = current_row, column = 5)
  op4 = sheet_obj.cell(row = current_row, column = 6)
  ans = sheet_obj.cell(row = current_row, column = 7)
  expl = sheet_obj.cell(row = current_row, column = 8)
  questions.close()

  question = ques.value
  options = [op1.value, op2.value, op3.value, op4.value]
  answer = ans.value
  explanation = expl.value
  telegram_poll_data = {}
  if(explanation is None):
    telegram_poll_data = {
      'chat_id': chat_id,
      'options':  options,
      'question': question,
      'type': "quiz",
      'correct_option_id': answer,
    }
  else:
    telegram_poll_data = {
        'chat_id': chat_id,
        'options':  options,
        'question': question,
        'type': "quiz",
        'correct_option_id': answer,
        'explanation':explanation
    }

  response = requests.post(telegram_poll_url, json=telegram_poll_data)
  if (response.status_code == 200):
    update_json(current_row+1,date_to_post)
    questions_left = total_ques-current_row
    if(questions_left<10):
      telegram_message_data = {
        'chat_id': my_chat_id,
        'text': "Running out of questions. "+str(questions_left)+" left."
      }
      response = requests.post(telegram_message_url, json=telegram_message_data)
  else:
    telegram_message_data = {
      'chat_id': my_chat_id,
      'text': "Failed to send todays' quiz."
    }
    response = requests.post(telegram_message_url, json=telegram_message_data)
def update_json(last_row,date_to_post):
  dictionary ={
    "last_row" : last_row,
    "date" : datetime.strftime(date_to_post,'%Y-%m-%d %H:%M:%S')
  }

  json_object = json.dumps(dictionary)
  with open("data.json", "w") as outfile:
      outfile.write(json_object)
      outfile.close()

def main():
  while(True):
    date_to_post_str = ""
    with open('data.json', 'r') as openfile:
      json_object = json.load(openfile)
      date_to_post_str = json_object["date"]
    date_to_post = datetime.strptime(date_to_post_str, '%Y-%m-%d %H:%M:%S')

    current_date = datetime.now()

    if(date_to_post.day == current_date.day and date_to_post.hour == current_date.hour):
      date_to_post = date_to_post + timedelta(days = 1)
      sendPoll(date_to_post)
      difference = date_to_post-current_date
      s = difference.total_seconds()
      time.sleep(s-60)

x = threading.Thread(target=main)
x.start()

bot = telebot.TeleBot(API_KEY)

@bot.message_handler(commands=['Start','start'])
def start(message):
  uid = message.chat.id
  if(str(uid) != str(my_chat_id)):
    print(uid,my_chat_id)
    bot.send_message(uid, "Only the admin can use this bot.")
    return
  markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard = True)
  itembtn1 = telebot.types.KeyboardButton('Add Question')
  itembtn2 = telebot.types.KeyboardButton('Remaining Questions')
  itembtn3 = telebot.types.KeyboardButton('Next Quiz')
  markup.add(itembtn1, itembtn2, itembtn3)
  bot.send_message(my_chat_id, "What would you like to do:", reply_markup=markup)

@bot.message_handler(func=lambda m: True)
def message_handler(message):
  try:
    uid = message.chat.id
    if(str(uid) != str(my_chat_id)):
      print(uid,my_chat_id)
      bot.send_message(uid, "Only the admin can use this bot.")
      return
    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard = True)
    itembtn1 = telebot.types.KeyboardButton('Add Question')
    itembtn2 = telebot.types.KeyboardButton('Remaining Questions')
    itembtn3 = telebot.types.KeyboardButton('Next Quiz')
    markup.add(itembtn1, itembtn2, itembtn3)
    if(message.text=="Add Question"):
      markupaq = telebot.types.ForceReply(selective=True)
      bot.send_message(my_chat_id, "Ok send the question:", reply_markup=markupaq)

    elif(message.text=="Remaining Questions"):
      questions = openpyxl.load_workbook("questions.xlsx")
      sheet_obj = questions.active
      total_ques = sheet_obj.max_row
      with open('data.json', 'r') as openfile:
        json_object = json.load(openfile)
        current_row = json_object["last_row"]-1
        bot.send_message(my_chat_id, "There are "+str(total_ques - current_row)+" Questions left.", reply_markup=markup)

    elif(message.text=="Next Quiz"):
      sendPollToAdmin()

    elif(message.reply_to_message != None):
      if(message.reply_to_message.json['text']=="Ok send the question:"):
        question_data = message.text
        question_list = question_data.split('\n')
        if(5<len(question_list)<8):
          question = question_list[0]
          ans1 = question_list[1]
          ans2 = question_list[2]
          ans3 = question_list[3]
          ans4 = question_list[4]
          cor_ans = question_list[5]
          expl = None
          if(len(question_list) == 7):
            expl = question_list[6]
          if(0<int(cor_ans)<5):
            questions = openpyxl.load_workbook("questions.xlsx")
            sheet_obj = questions.active
            max_row = sheet_obj.max_row
            sheet_obj["A"+str(max_row+1)] = max_row
            sheet_obj["B"+str(max_row+1)] = question
            sheet_obj["C"+str(max_row+1)] = ans1
            sheet_obj["D"+str(max_row+1)] = ans2
            sheet_obj["E"+str(max_row+1)] = ans3
            sheet_obj["F"+str(max_row+1)] = ans4
            sheet_obj["G"+str(max_row+1)] = int(cor_ans)-1
            sheet_obj["H"+str(max_row+1)] = expl
            questions.save(filename="questions.xlsx")
            questions.close()
            bot.send_message(my_chat_id, "Quiz has been saved.", reply_markup=markup)
          else:
            bot.send_message(my_chat_id, "Index for correct answer must be between 1 - 4.")
        else:
          bot.send_message(my_chat_id, "Something is missing. Try again")
  except:
     bot.send_message(my_chat_id, "Incorrect input.")

def sendPollToAdmin():
  questions = openpyxl.load_workbook("questions.xlsx")
  sheet_obj = questions.active
  total_ques = sheet_obj.max_row

  current_row = 2
  with open('data.json', 'r') as openfile:
    json_object = json.load(openfile)
    current_row = json_object["last_row"]

  ques = sheet_obj.cell(row = current_row, column = 2)
  op1 = sheet_obj.cell(row = current_row, column = 3)
  op2 = sheet_obj.cell(row = current_row, column = 4)
  op3 = sheet_obj.cell(row = current_row, column = 5)
  op4 = sheet_obj.cell(row = current_row, column = 6)
  ans = sheet_obj.cell(row = current_row, column = 7)
  expl = sheet_obj.cell(row = current_row, column = 8)
  questions.close()

  question = ques.value
  options = [op1.value, op2.value, op3.value, op4.value]
  answer = ans.value
  explanation = expl.value
  telegram_poll_data = {}
  if(explanation is None):
    telegram_poll_data = {
      'chat_id': my_chat_id,
      'options':  options,
      'question': question,
      'type': "quiz",
      'correct_option_id': answer,
    }
  else:
    telegram_poll_data = {
        'chat_id': my_chat_id,
        'options':  options,
        'question': question,
        'type': "quiz",
        'correct_option_id': answer,
        'explanation':explanation
    }

  response = requests.post(telegram_poll_url, json=telegram_poll_data)
  if (response.status_code != 200):
    telegram_message_data = {
      'chat_id': my_chat_id,
      'text': "Failed to send next quiz."
    }
    response = requests.post(telegram_message_url, json=telegram_message_data)

bot.polling()
